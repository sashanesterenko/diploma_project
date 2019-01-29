import requests
import json
from bs4 import BeautifulSoup as bs
import csv
import pandas as pd
from openpyxl import Workbook
from openpyxl.compat import range
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
from collections import OrderedDict

notice_url = 'http://zakupki.gov.ru/epz/order/notice/printForm/view.html?printFormId=81135264'

def get_html(url, headers):
    try:
        result = requests.get(url, headers=headers)
        result.raise_for_status()  # if result.status_code >= 400
        return result.text
    except (requests.RequestException, ValueError):
        return False


def parse_notice_info(souped_html):
    all_tr = souped_html.find_all('tr')
    param_value_dict = {}
    for tr in all_tr:
        param = tr.find('p', class_='parameter')
        param_value = tr.find('p', class_='parameterValue')
        if param and param_value:
            param = param.text
            param_value = param_value.text
            param_value_pair = {param:param_value}
            param_value_dict.update(param_value_pair)
    needed_keys = ['Номер извещения',
            'Способ определения поставщика (подрядчика, исполнителя)', 
            'Организация, осуществляющая размещение',
            'Почтовый адрес',
            'Адрес электронной площадки в информационно-телекоммуникационной сети «Интернет»', 
            'Начальная (максимальная) цена контракта',
            'Дата и время подписания печатной формы извещения (соответствует дате направления на контроль по ч.5 ст.99 Закона 44-ФЗ либо дате размещения в ЕИС, в случае отсутствия контроля, по местному времени организации, осуществляющей размещение)',
            'Дата и время окончания подачи заявок', 
            'Дата проведения аукциона в электронной форме']
    needed_param_value_dict = {}
    for key in param_value_dict.keys():
        if key in needed_keys:
            # print(needed_param_value_dict)
            needed_param_value_dict.update({key: param_value_dict[key]})
            # print(needed_param_value_dict)
        else:
            pass
    needed_param_value_dict = dict(sorted(needed_param_value_dict.items(), key=lambda pair:needed_keys.index(pair[0])))
    # print(needed_param_value_dict)
    return needed_param_value_dict

def parse_notice_goods(souped_html):
    table_body = souped_html.find_all('table', class_='table font9')
    for tb in table_body:
        tbody_tr = tb.find_all('tr', class_=False)  
        # print(tbody_tr)           
    table_param_value_dict = {}
    table_param_value_list = []
    skiped_header = 'Характеристики товара, работы, услуги'
    for tr in tbody_tr:
        try_th = tr.find_all('th')
        if try_th:
            tbody_th = [every_th.text for every_th in try_th if every_th.text != skiped_header]
            # for every_th in try_th:
            #     every_th_text = every_th.text
            #     if every_th_text != 'Характеристики товара, работы, услуги':
            #         # print(every_th_text)
            #         tbody_th.append(every_th_text)
            # print(tbody_th)
        else:
            try_td = tr.find_all('td', style=False)
            # print(try_td)
            if try_td: 
                tbody_td = [every_td.text for every_td in try_td if every_td.text != '']           
                # for every_td in try_td:
                #     every_td_text = every_td.text
                #     # print(every_td_text)
                #     if every_td_text != '':
                #         tbody_td.append(every_td_text)
                # print(tbody_td)
                table_param_value_dict = dict(zip(tbody_th, tbody_td))
                    # print(table_param_value_dict)
                table_param_value_list.append(table_param_value_dict)

    needed_table_keys = {'Наименование товара, работы, услуги по КТРУ',
                        'Количество', 
                        'Цена за ед.изм.'}
    needed_table_param_value_list = []
    for a_dict in table_param_value_list:
        needed_table_dict = {}
        for key in a_dict.keys():
            if key in needed_table_keys:
                needed_table_dict.update({key: a_dict[key]})
        needed_table_param_value_list.append(needed_table_dict)
    # print(needed_table_param_value_list)
    return needed_table_param_value_list
    # print(table_param_value_list)



def get_relevant_info():
    html = get_html(notice_url, {'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_13_6) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/70.0.3538.110 Safari/537.36'})
    if not html:
        return
    soup = bs(html, 'html.parser')
    needed_param_value_dict = parse_notice_info(soup)
    needed_table_param_value_list = parse_notice_goods(soup)
    return needed_param_value_dict, needed_table_param_value_list


def excel_export(needed_param_value_dict, needed_table_param_value_list):
    fields = ['Способ определения поставщика (подрядчика, исполнителя)', 
            'Организация, осуществляющая размещение',
            'Почтовый адрес',
            'Наименование товара, работы, услуги по КТРУ',
            'Количество', 
            'Цена за ед.изм.',                
            'Адрес электронной площадки в информационно-телекоммуникационной сети «Интернет»', 
            'Начальная (максимальная) цена контракта', 
            'Дата и время подписания печатной формы извещения (соответствует дате направления на контроль по ч.5 ст.99 Закона 44-ФЗ либо дате размещения в ЕИС, в случае отсутствия контроля, по местному времени организации, осуществляющей размещение)',
            'Дата и время окончания подачи заявок', 
            'Дата проведения аукциона в электронной форме']
    column_range = [3, 4, 5, 12, 13, 14, 15, 16]
    column_range_goods = [6, 7, 8]
    wb = load_workbook(filename='test_table.xlsx')
    sheet = wb.active
    rows_index = 0
    for row in sheet.iter_rows(min_col=1, max_col=1):
        for cell in row:
            if cell.value != None:
                rows_index = cell.value
        
    index_number = rows_index + 1
    sheet.cell(column=1, row=sheet.max_row+1, value=index_number)
        
    column_range_index = 0
    for key in needed_param_value_dict.keys():
        if key in fields:
            try:
                sheet.cell(column=column_range[column_range_index], row=sheet.max_row, value=needed_param_value_dict[key])
                column_range_index = column_range_index + 1
            except (IndexError, KeyError):
                pass

    row_index = sheet.max_row
    for a_dict in needed_table_param_value_list:
        column_range_index = 0
        for key in a_dict.keys():
            if key in fields:
                try:
                    sheet.cell(column=column_range_goods[column_range_index], row=row_index, value=a_dict[key])
                    column_range_index += 1
                except (IndexError):
                    pass
        row_index +=1 

    wb.save(filename='test_table.xlsx')




def bd_export(needed_param_value_dict, needed_table_param_value_list):
   
    json_export = {'zakupkigov_id': needed_param_value_dict['Номер извещения'], 
                'vendee_definition_method': needed_param_value_dict['Способ определения поставщика (подрядчика, исполнителя)'],
                'vendee_info': {'postal_address': needed_param_value_dict['Почтовый адрес'],
                        'vendee': needed_param_value_dict['Организация, осуществляющая размещение'],
                        'notice_id': needed_param_value_dict['Номер извещения']}}
    items = []
    for item in needed_table_param_value_list:
        items.append({'name': item['Наименование товара, работы, услуги по КТРУ'],
                    'quantity': item['Количество'],
                    'price': item['Цена за ед.изм.']})
    json_export.update({'goods': items})
    json_export.update({'tender_system': needed_param_value_dict['Адрес электронной площадки в информационно-телекоммуникационной сети «Интернет»'], 
                    'start_max_price': needed_param_value_dict['Начальная (максимальная) цена контракта'], 
                    'application_dates_end': needed_param_value_dict['Дата и время окончания подачи заявок'],
                    'e_auction_date': needed_param_value_dict['Дата проведения аукциона в электронной форме'],
                    'date_printed_notice': needed_param_value_dict['Дата и время подписания печатной формы извещения (соответствует дате направления на контроль по ч.5 ст.99 Закона 44-ФЗ либо дате размещения в ЕИС, в случае отсутствия контроля, по местному времени организации, осуществляющей размещение)']})



    

    json_export = json.dumps(json_export, ensure_ascii=False)
    return json_export


if __name__ == '__main__':

    needed_param_value_dict, needed_table_param_value_list = get_relevant_info()

    excel_export(needed_param_value_dict, needed_table_param_value_list)

    json_export = bd_export(needed_param_value_dict, needed_table_param_value_list)

    print(json_export)




